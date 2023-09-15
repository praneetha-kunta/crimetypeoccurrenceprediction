from django.db.models import Count
from django.db.models import Q
from django.shortcuts import render, redirect, get_object_or_404
import datetime
import openpyxl


# Create your views here.
from Remote_User.models import ClientRegister_Model,Crime_details,Crime_type,detection_ratio,detection_accuracy

def login(request):


    if request.method == "POST" and 'submit1' in request.POST:

        username = request.POST.get('username')
        password = request.POST.get('password')
        try:
            enter = ClientRegister_Model.objects.get(username=username,password=password)
            request.session["userid"] = enter.id

            return redirect('Add_DataSet_Details')
        except:
            pass

    return render(request,'RUser/login.html')

def Add_DataSet_Details(request):
    if "GET" == request.method:
        return render(request, 'RUser/Add_DataSet_Details.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting all sheets
        sheets = wb.sheetnames
        #print(sheets)
        # getting a particular sheet
        worksheet = wb["Sheet1"]
        #print(worksheet)
        # getting active sheet
        active_sheet = wb.active
        #print(active_sheet)
        # reading a cell
        #print(worksheet["A1"].value)
        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
                #print(cell.value)
            excel_data.append(row_data)
            Crime_details.objects.all().delete()
            Crime_type.objects.all().delete()
    for r in range(1, active_sheet.max_row + 1):
        Crime_details.objects.create(
        INCIDENT_NUMBER=active_sheet.cell(r, 1).value,
        OFFENSE_CODE=active_sheet.cell(r, 2).value,
        OFFENSE_CODE_GROUP=active_sheet.cell(r, 3).value,
        OFFENSE_DESCRIPTION=active_sheet.cell(r, 4).value,
        DISTRICT=active_sheet.cell(r, 5).value,
        REPORTING_AREA=active_sheet.cell(r, 6).value,
        OCCURRED_ON_DATE=active_sheet.cell(r, 7).value,
        YEAR=active_sheet.cell(r, 8).value,
        MONTH=active_sheet.cell(r, 9).value,
        DAY_OF_WEEK=active_sheet.cell(r, 10).value,
        Hour=active_sheet.cell(r, 11).value,
        UCR_PART=active_sheet.cell(r, 12).value,
        STREET=active_sheet.cell(r, 13).value,
        Lat=active_sheet.cell(r, 14).value,
        Long1=active_sheet.cell(r, 15).value,
        Location=active_sheet.cell(r, 16).value
        )

    return render(request, 'RUser/Add_DataSet_Details.html', {"excel_data": excel_data})


def Register1(request):

    if request.method == "POST":
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        phoneno = request.POST.get('phoneno')
        country = request.POST.get('country')
        state = request.POST.get('state')
        city = request.POST.get('city')
        ClientRegister_Model.objects.create(username=username, email=email, password=password, phoneno=phoneno,
                                            country=country, state=state, city=city)

        return render(request, 'RUser/Register1.html')
    else:
        return render(request,'RUser/Register1.html')

def ViewYourProfile(request):
    userid = request.session['userid']
    obj = ClientRegister_Model.objects.get(id= userid)
    return render(request,'RUser/ViewYourProfile.html',{'object':obj})


def Search_DataSets(request):
    if request.method == "POST":
        kword = request.POST.get('keyword')
        if request.method == "POST":
            kword = request.POST.get('keyword')

        obj = Crime_type.objects.all().filter(Q(INCIDENT_NUMBER__contains=kword))

        return render(request, 'RUser/Search_DataSets.html',{'objs': obj})
    return render(request, 'RUser/Search_DataSets.html')



